# ГЛАВНЫЙ ФАЙЛ
from openpyxl import load_workbook
from distances import distances
from distances import print_array

def euclid_distance(c1, c2):
    distance = 0
    for i in range(4):
        distance += (c1[i] - c2[i])**2
    return distance**(1/2)

# Читаем эксель-файл
wb = load_workbook(filename = 'irisy.xlsx')
l = wb['l']

iris_coordinates = []
iris_classes = []

# Заполняем списки с координатами и классами ирисов
for i in range(2,122):
    length1 = l.cell(row=i,column=1).value
    width1 = l.cell(row=i,column=2).value
    length2 = l.cell(row=i,column=3).value
    width2 = l.cell(row=i,column=4).value
    iris_class = l.cell(row=i,column=5).value
    coords = [length1, width1, length2, width2, iris_class]
    iris_coordinates.append(coords)

# Считаем расстояние между ирисами методом Евклида
iris_size = len(iris_coordinates)
dis = [0]*(iris_size)
i = 0
iris_classes = []
for c1 in iris_coordinates:
    dist = []
    for c2 in iris_coordinates:
        if(iris_coordinates.index(c1)==iris_coordinates.index(c2)):
            dist.append(1000)
        else:
            dist.append(euclid_distance(c1,c2))
    iris_classes.append(c1[4])
    dis[i] = dist
    i += 1
# Методом агломерационной кластеризации находим основные кластеры и их расстояние между друг другом
dis = distances(dis,iris_classes,3)
print_array(dis, iris_classes)