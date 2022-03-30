from openpyxl import load_workbook
wb = load_workbook(filename = 'irisy.xlsx')
l = wb['l']
r = wb['r']
for i in range(2,32):
    for j in range(2,122):
        l1 = abs(l.cell(row=j,column=3).value - r.cell(row=i,column=3).value)
        l2 = abs(l.cell(row=j,column=4).value - r.cell(row=i,column=4).value)
        l.cell(row=j,column=7).value = l1 + l2
    min = l.cell(row=2,column=7).value
    n_min = 2
    for j in range(3,121):
        if l.cell(row=j,column=7).value < min:
            min = l.cell(row=j,column=7).value
            n_min=j
    l.cell(row=1,column=7).value = min
    l.cell(row=1,column=9).value = n_min
    r.cell(row=i,column=7).value = l.cell(row=n_min, column=5).value
r.cell(row=1,column=7).value = "Найденный класс"
wb.save('irisy.xlsx')