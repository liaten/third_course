import re
from tokenize import Double
from warnings import catch_warnings
from openpyxl import load_workbook

print("Программа возвращает вероятность события относительно 4 признаков.")

# объявление переменных
y = -1 # да - (1), нет - (0)
obl = 0 # облачность (1) - солнечно (2) - пасмурно (3) - дождь
temp = 0 # температура (1) - жарко (2) - тепло (3) - прохладно
vlazh = 0 # влажность (1) - высокая (2) - нормальная
vet = 0 # ветер (1) - умеренный (2) - сильный

print("Состоится ли матч?")
# считывание ввода пользователя
while 1: # 1 приводится к True
    try:
        y = int(input("(0) - нет (1) - да\n:"))
        if y in range(0,2):
            break
    except ValueError: # неверное значение ввода
        continue

print("Выберите признаки:")
while 1:
    try:
        obl = int(input("Облачность\n(1) - солнечно (2) - пасмурно (3) - дождь\n:"))
        if(obl in range(1,4)):
            break
    except ValueError:
        continue
while 1:
    try:
        temp = int(input("Температура\n(1) - жарко (2) - тепло (3) - прохладно\n:"))
        if(temp in range(1,4)):
            break
    except ValueError:
        continue
while 1:
    try:
        vlazh = int(input("Влажность\n(1) - высокая (2) - нормальная\n:"))
        if(vlazh in range(1,3)):
            break
    except ValueError:
        continue
while 1:
    try:
        vet = int(input("Ветер\n(1) - умеренный (2) - сильный\n:"))
        if(vet in range(1,3)):
            break
    except ValueError:
        continue

wb = load_workbook(filename = 'bayes.xlsx')
init = wb['вероятности']
print(init.cell(row=1,column=1).value)

# объявление переменных вероятностей событий
# что умножать?
p_y = 0
p_x1_cond = 0 # вероятность при условии (да/нет)
p_x2_cond = 0
p_x3_cond = 0
p_x4_cond = 0

# на что делить?
p_x1 = 0
p_x2 = 0
p_x3 = 0
p_x4 = 0

# присваивание p(y)
match(y):
    case 0:
        p_y = float(init.cell(row=19,column=3).value)
    case 1:
        p_y = float(init.cell(row=19,column=2).value)

# присваивание вероятностей событий из таблицы (p(x) p(x_cond))
p_x1_cond = float(init.cell(row=1+obl,column=5-y).value)
p_x1 = float(init.cell(row=1+obl,column=6).value)

p_x2_cond = float(init.cell(row=6+temp,column=5-y).value)
p_x2 = float(init.cell(row=6+temp,column=6).value)

p_x3_cond = float(init.cell(row=11+vlazh,column=5-y).value)
p_x3 = float(init.cell(row=11+vlazh,column=6).value)

p_x4_cond = float(init.cell(row=15+vet,column=5-y).value)
p_x4 = float(init.cell(row=15+vet,column=6).value)

result = (p_x1_cond * p_x2_cond * p_x3_cond * p_x4_cond * p_y) / (p_x1 * p_x2 * p_x3 * p_x4)

# сборка строки вывода
result_s = "Вероятность, что матч "
match(y):
    case 0:
        result_s += "не "
result_s += "состоится = "+ ("%.5f" % result)

# вывод результата
print(result_s)

wb.close() # закрытие рабочего пространства