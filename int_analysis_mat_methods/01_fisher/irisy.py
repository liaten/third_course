from openpyxl import load_workbook
wb = load_workbook(filename = 'irisy.xlsx')
# l, r - рабочие пространства
l = wb['l']
r = wb['r']
r.cell(row=1,column=7).value = "Найденный класс"
# i - значение строки искомого растения
for i in range(2,32):
    # j - значение строки растения из базы
    for j in range(2,122):
        l1 = abs(l.cell(row=j,column=3).value - r.cell(row=i,column=3).value)
        l2 = abs(l.cell(row=j,column=4).value - r.cell(row=i,column=4).value)
        l.cell(row=j,column=7).value = l1 + l2
    # объявление первого числа минимальным и сохранение его
    min = l.cell(row=2,column=7).value
    n_min = 2
    # нахождение минимального числа и его позиции
    for j in range(3,121):
        if l.cell(row=j,column=7).value < min:
            min = l.cell(row=j,column=7).value
            n_min=j
    # запись в excel-файл
    l.cell(row=1,column=7).value = min
    l.cell(row=1,column=9).value = n_min
    r.cell(row=i,column=7).value = l.cell(row=n_min, column=5).value
# сохраняем и закрываем файл
wb.save('irisy.xlsx')
wb.close()